import streamlit as st
import requests
import json
import os
import re
import anthropic
import pandas as pd
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import datetime
import base64
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Get API keys from environment variables
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')
SEARCH_API_KEY = os.getenv('SEARCH_API_KEY')

# Validate API keys
if not ANTHROPIC_API_KEY or not SEARCH_API_KEY:
    raise ValueError("Missing required API keys in .env file")

def extract_file_ids_from_folder(folder_url):
    """Extract file IDs from Google Drive folder"""
    try:
        folder_id = folder_url.split('/')[-1]
        files_url = f"https://drive.google.com/drive/folders/{folder_id}"
        response = requests.get(files_url)
        
        pattern = r"https://drive\.google\.com/file/d/([a-zA-Z0-9_-]+)"
        file_ids = list(set(re.findall(pattern, response.text)))
        
        image_urls = []
        for file_id in file_ids:
            direct_link = f"https://drive.google.com/uc?id={file_id}"
            image_urls.append({
                'id': file_id,
                'url': direct_link,
                'name': f"image_{file_id}.jpg"
            })
        
        return image_urls
    except Exception as e:
        st.error(f"Error extracting file IDs: {str(e)}")
        return []

def get_anthropic_analysis(json_data):
    """Get analysis from Anthropic's Claude API"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    
    prompt = f"""Please analyze the following product search results and provide a structured summary 
    with clear headings and bullet points. Focus on:
    1. Price Range Analysis
    2. Common Product Types
    3. Notable Patterns or Trends
    4. Key Findings or Recommendations
    5. limit to maximum 20 words

    Here's the data:
    {json.dumps(json_data, indent=2)}"""
    
    try:
        message = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        content = message.content
        return content[0].text if not isinstance(content, str) else content
    except Exception as e:
        st.error(f"Error in Anthropic API call: {str(e)}")
        return "Error in generating analysis"

def create_excel_report(results, output_file):
    """Creates an Excel report with images and analyses"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set up headers
    headers = ['Image', 'File Name', 'Analysis']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    row_height_multiplier = 15
    
    # Add data and images
    for row_idx, result in enumerate(results, 2):
        try:
            # Add image
            if os.path.exists(result['temp_image_path']):
                img = XLImage(result['temp_image_path'])
                img.width = 200
                img.height = 200
                ws.add_image(img, f'A{row_idx}')
            
            # Add filename
            ws.cell(row=row_idx, column=2, value=result['name'])
            
            # Add analysis
            analysis_cell = ws.cell(row=row_idx, column=3, value=result['analysis'])
            analysis_cell.alignment = openpyxl.styles.Alignment(
                wrap_text=True,
                vertical='top'
            )
            
            # Auto-adjust row height
            text_lines = len(result['analysis'].split('\n'))
            ws.row_dimensions[row_idx].height = max(200, text_lines * row_height_multiplier)
            
        except Exception as e:
            print(f"Error adding row {row_idx} to Excel: {str(e)}")
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    
    # Add borders
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(results)+1, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    try:
        wb.save(output_file)
        print(f"Excel report saved as: {os.path.abspath(output_file)}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        return False

def search_google_lens(image_url):
    """Search using Google Lens API"""
    api_url = "https://www.searchapi.io/api/v1/search"
    params = {
        "engine": "google_lens",
        "search_type": "all",
        "url": image_url,
        "api_key": SEARCH_API_KEY
    }
    
    try:
        response = requests.get(api_url, params=params)
        response.raise_for_status()
        response_json = response.json()
        visual_matches = response_json.get("visual_matches", [])[:15]
        
        filtered_results = []
        for match in visual_matches:
            filtered_entry = {
                "position": match.get("position"),
                "title": match.get("title"),
                "source": match.get("source"),
                "price": match.get("price", "N/A"),
                "extracted_price": match.get("extracted_price", 0.0),
                "currency": match.get("currency", "N/A")
            }
            filtered_results.append(filtered_entry)
        
        return filtered_results
    except Exception as e:
        st.error(f"Error in Google Lens search: {str(e)}")
        return []

def create_unique_filename(base_name="report"):
    """Create unique filename with timestamp"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{timestamp}.xlsx"

def download_link(file_path, file_name):
    """Generate download link for Excel file"""
    with open(file_path, 'rb') as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">Download Excel Report</a>'
    return href

def main():
    st.set_page_config(page_title="Estate Sale AI Appraisal App ", page_icon="🔍", layout="wide")
    
    st.title("🔍 EstateGenius AI")
    st.markdown("---")

    with st.sidebar:
        st.header("About")
        st.write("""
        The Estate Sale Game-Changer: No more item-by-item analysis. 
        Our AI processes hundreds of photos in one go, delivering 
        comprehensive Excel reports with detailed valuations.
        Built for efficiency, trusted by professionals.
        """)

    folder_url = st.text_input("Enter Google Drive Folder URL", 
                              placeholder="https://drive.google.com/drive/folders/...")

    if st.button("Process Images", type="primary"):
        if not folder_url:
            st.error("Please enter a valid folder URL")
        else:
            try:
                with st.spinner("Processing..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    status_text.text("Extracting images from folder...")
                    images = extract_file_ids_from_folder(folder_url)
                    progress_bar.progress(20)

                    if not images:
                        st.error("No images found in the folder")
                    else:
                        results = []
                        temp_files = set()

                        for i, image in enumerate(images):
                            progress = 20 + (60 * (i + 1) / len(images))
                            progress_bar.progress(int(progress))
                            status_text.text(f"Processing image {i+1}/{len(images)}")

                            temp_path = f"temp_image_{image['id']}.png"
                            temp_files.add(temp_path)

                            try:
                                response = requests.get(image['url'])
                                if response.status_code == 200:
                                    img = Image.open(BytesIO(response.content))
                                    img.convert('RGB').save(temp_path, 'PNG')

                                    lens_results = search_google_lens(image['url'])
                                    if lens_results:
                                        analysis = get_anthropic_analysis(lens_results)
                                        results.append({
                                            'name': image['name'],
                                            'temp_image_path': temp_path,
                                            'analysis': analysis
                                        })
                            except Exception as e:
                                st.error(f"Error processing image: {str(e)}")

                        if results:
                            status_text.text("Generating Excel report...")
                            progress_bar.progress(90)
                            
                            output_file = create_unique_filename()
                            success = create_excel_report(results, output_file)

                            if success:
                                progress_bar.progress(100)
                                status_text.text("Processing complete!")
                                st.markdown(download_link(output_file, output_file), unsafe_allow_html=True)
                                st.success("Report generated successfully!")
                            else:
                                st.error("Failed to create Excel report")

                        for temp_file in temp_files:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)

            except Exception as e:
                st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    try:
        if not ANTHROPIC_API_KEY or not SEARCH_API_KEY:
            st.error("Missing required API keys. Please check your .env file.")
        else:
            main()
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")